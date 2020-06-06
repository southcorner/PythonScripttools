import bs4 as bs 
import sys 
import schedule 
import time 
from PyQt5.QtWebEngineWidgets import QWebEnginePage 
from PyQt5.QtWidgets import QApplication 
from PyQt5.QtCore import QUrl
import openpyxl

  
#import winsound 
frequency = 2500  # Set Frequency To 2500 Hertz 
duration = 1000  # Set Duration To 1000 ms == 1 second 

path = "YOUR PATH TO XLSX HERE" 
wb_obj = openpyxl.load_workbook(path.strip())
# from the active attribute 
sheet_obj = wb_obj.active
max_column=sheet_obj.max_column
max_row=sheet_obj.max_row
  
class Page(QWebEnginePage): 
  
    def __init__(self, url): 
        self.app = QApplication(sys.argv) 
        QWebEnginePage.__init__(self) 
        self.html = '' 
        self.loadFinished.connect(self._on_load_finished) 
        self.load(QUrl(url)) 
        self.app.exec_() 
  
    def _on_load_finished(self): 
        self.html = self.toHtml(self.Callable) 
  
    def Callable(self, html_str): 
        self.html = html_str 
        self.app.quit() 
  
def exact_url(url): 
    current_url = "https://www.amazon.in/dp/"+url 
    return current_url 
      
  
def mainprogram():
    i = 1
    url_cell = sheet_obj.cell(row=i,column=1) #IF XLSX IS ALTERED AND BO COLUMN CHANGES CHANGE COLUMN VALUE
    url = url_cell.value
    while url_cell.value:
        
        exacturl = exact_url(url)
        # main url to extract data 
        page = Page(exacturl) 
        soup = bs.BeautifulSoup(page.html, 'html.parser') 
        js_test = soup.find('span', id ='priceblock_ourprice')
        if js_test is None: 
            js_test = soup.find('span', id ='priceblock_dealprice')
        if js_test is None:
            print("Invalid URL")
        str = "" 
        for line in js_test.stripped_strings : 
            str = line
  
        # convert to integer 
        str = str.replace(", ", "") 
        current_price = str
        print(current_price)
        price_cell = sheet_obj.cell(row=i,column=2)
        price_cell.value = current_price
        wb_obj.save("BO_list.xlsx")
        i += 1
        url_cell = sheet_obj.cell(row=i,column=1)
        url = url_cell.value
      
def job():     
    mainprogram() 
  
 
schedule.every(1).minutes.do(job) 
  
while True: 
    schedule.run_pending() 
    time.sleep(1)